#!/usr/bin/env python3
import io
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

BASE = "https://www.reb.or.kr"
LIST_URL = BASE + "/reb/na/ntt/selectNttList.do?mi=9565&bbsId=1154"
RONE_REPORT_URL = BASE + "/r-one/portal/bbs/rpt/searchBulletinPage.do"
OUT = Path("real-estate-pulse-data.json")
REGIONS = ["서울","부산","대구","인천","광주","대전","울산","세종","경기","강원","충북","충남","전북","전남","경북","경남","제주"]
UA = {"User-Agent": "Mozilla/5.0 (compatible; NotionRealEstatePulse/1.2; +https://github.com/Ohoredingdong/Notion)"}


def get(url, **kwargs):
    r = requests.get(url, headers=UA, timeout=35, **kwargs)
    r.raise_for_status()
    return r


def latest_release():
    candidates = []
    for page in range(1, 4):
        url = LIST_URL + (f"&currPage={page}" if page > 1 else "")
        soup = BeautifulSoup(get(url).text, "html.parser")
        for a in soup.find_all("a"):
            title = " ".join(a.stripped_strings)
            m = re.search(r"주간아파트가격동향\((\d{8})기준\)", title)
            if not m:
                continue
            href = a.get("href", "")
            if not href:
                continue
            if href.startswith("javascript:"):
                n = re.search(r"nttSn\D+(\d+)", href)
                if n:
                    href = f"/reb/na/ntt/selectNttInfo.do?mi=9565&bbsId=1154&nttSn={n.group(1)}"
                else:
                    continue
            candidates.append((m.group(1), urljoin(BASE, href)))
    if not candidates:
        raise RuntimeError("REB 주간아파트가격동향 최신 게시물을 찾지 못했습니다.")
    return max(candidates, key=lambda x: x[0])


def urls_from_html(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for tag in soup.find_all(True):
        text = " ".join(tag.stripped_strings)
        for key in ("href", "src", "data-url", "data-file", "onclick"):
            raw = tag.get(key)
            if not raw:
                continue
            vals = raw if isinstance(raw, list) else [raw]
            for value in vals:
                value = str(value).replace("&amp;", "&")
                if value.startswith("/") or value.startswith("http"):
                    out.append((urljoin(base_url, value), text))
                for match in re.findall(r"['\"]([^'\"]+)['\"]", value):
                    if match.startswith("/") or match.startswith("http"):
                        out.append((urljoin(base_url, match.replace("&amp;", "&")), text))
    for raw in re.findall(r"https?://[^'\"\s<>]+|/[^'\"\s<>]*(?:xlsx|xls|download|file)[^'\"\s<>]*", html, re.I):
        out.append((urljoin(base_url, raw.replace("&amp;", "&")), ""))
    seen = set()
    unique = []
    for url, text in out:
        if url not in seen:
            seen.add(url)
            unique.append((url, text))
    return unique


def is_excel_response(r):
    ct = r.headers.get("content-type", "").lower()
    cd = r.headers.get("content-disposition", "").lower()
    head = r.content[:8]
    return (
        head.startswith(b"PK\x03\x04")
        or head.startswith(b"\xD0\xCF\x11\xE0")
        or "spreadsheet" in ct
        or "ms-excel" in ct
        or ".xlsx" in cd
        or ".xls" in cd
    )


def discover_rone_excel(asof_raw):
    """Best-effort discovery of the official R-ONE weekly Excel.

    R-ONE's report list is partly dynamic, so try several common search parameter
    names and then follow report/detail links. If the site changes, PDF parsing is
    the fallback and the existing JSON is never replaced with partial data.
    """
    page_responses = []
    searches = [
        {},
        {"searchKeyword": "전국주택가격동향조사"},
        {"searchWord": "전국주택가격동향조사"},
        {"searchWrd": "전국주택가격동향조사"},
        {"searchText": "전국주택가격동향조사"},
        {"searchKeyword": "주간"},
    ]
    for params in searches:
        try:
            r = get(RONE_REPORT_URL, params=params)
            page_responses.append((r.url, r.text))
        except Exception as e:
            print("R-ONE report page request failed:", params, e, file=sys.stderr)

    detail_urls = []
    excel_urls = []
    for page_url, html in page_responses:
        for url, text in urls_from_html(html, page_url):
            blob = f"{url} {text}".lower()
            if ".xlsx" in blob or ".xls" in blob or "엑셀" in text:
                excel_urls.append(url)
            if "/r-one/portal/bbs/rpt/" in url and "searchBulletinPage.do" not in url:
                detail_urls.append(url)

    for detail_url in list(dict.fromkeys(detail_urls))[:40]:
        try:
            r = get(detail_url)
        except Exception:
            continue
        for url, text in urls_from_html(r.text, r.url):
            blob = f"{url} {text}".lower()
            if ".xlsx" in blob or ".xls" in blob or "엑셀" in text or "download" in blob or "file" in blob:
                excel_urls.append(url)

    # Prefer URLs that visibly contain the current YYYYMMDD, then weekly wording.
    excel_urls = list(dict.fromkeys(excel_urls))
    excel_urls.sort(key=lambda u: ((asof_raw in u), ("week" in u.lower() or "주간" in u), ("xlsx" in u.lower())), reverse=True)
    for url in excel_urls[:80]:
        try:
            r = get(url, allow_redirects=True)
        except Exception:
            continue
        if is_excel_response(r):
            print("R-ONE Excel candidate:", r.url)
            return r.url, r.content
    raise RuntimeError("R-ONE 공표보고서에서 주간 Excel 첨부파일을 자동 발견하지 못했습니다.")


def norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def date_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    s = norm(value)
    m = re.search(r"(20\d{2})[.\-/년]?(\d{1,2})[.\-/월]?(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}{int(m.group(2)):02d}{int(m.group(3)):02d}"
    return None


def numeric(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = norm(value).replace("%", "")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
        try:
            return float(s)
        except ValueError:
            return None
    return None


def excel_mode_values(excel_bytes, mode, asof_raw):
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    keyword = "매매" if mode == "sale" else "전세"
    candidates = {r: [] for r in REGIONS}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows())
        for row_index, row in enumerate(rows):
            region_cells = []
            for cell in row:
                n = norm(cell.value)
                if n in REGIONS:
                    region_cells.append((n, cell.column))
            if not region_cells:
                continue

            start = max(0, row_index - 24)
            context_cells = [norm(c.value) for rr in rows[start:row_index + 1] for c in rr if c.value is not None]
            context = norm(ws.title) + " " + " ".join(context_cells)
            if keyword not in context:
                continue

            # Find dated columns near the row. Prefer the exact official base date.
            dated_cols = {}
            for rr in rows[start:row_index + 1]:
                for c in rr:
                    dk = date_key(c.value)
                    if dk:
                        dated_cols[c.column] = dk

            for region, region_col in region_cells:
                current_col = None
                previous_col = None
                exact_cols = [col for col, dk in dated_cols.items() if dk == asof_raw]
                if exact_cols:
                    current_col = max(exact_cols)
                    older = [(dk, col) for col, dk in dated_cols.items() if dk < asof_raw]
                    if older:
                        previous_col = max(older)[1]
                elif dated_cols:
                    ordered = sorted((dk, col) for col, dk in dated_cols.items())
                    current_col = ordered[-1][1]
                    if len(ordered) >= 2:
                        previous_col = ordered[-2][1]

                current = None
                previous = None
                if current_col and current_col <= len(row):
                    current = numeric(row[current_col - 1].value)
                if previous_col and previous_col <= len(row):
                    previous = numeric(row[previous_col - 1].value)

                if current is None:
                    nums = [(c.column, numeric(c.value)) for c in row if c.column > region_col]
                    nums = [(col, val) for col, val in nums if val is not None and -10 <= val <= 10]
                    if nums:
                        current = nums[-1][1]
                        previous = nums[-2][1] if len(nums) >= 2 else None

                if current is None or not (-10 <= current <= 10):
                    continue
                score = 0
                if keyword in norm(ws.title): score += 8
                if "주간" in context: score += 6
                if "아파트" in context: score += 5
                if any(dk == asof_raw for dk in dated_cols.values()): score += 10
                if region_col <= 3: score += 2
                candidates[region].append((score, current, previous))

    values = {}
    previous = {}
    for region in REGIONS:
        if candidates[region]:
            score, cur, prev = max(candidates[region], key=lambda x: x[0])
            values[region] = cur
            previous[region] = prev
    return values, previous


def parse_excel(excel_bytes, asof_raw):
    sale, sale_prev = excel_mode_values(excel_bytes, "sale", asof_raw)
    lease, lease_prev = excel_mode_values(excel_bytes, "lease", asof_raw)
    missing_sale = [r for r in REGIONS if r not in sale]
    missing_lease = [r for r in REGIONS if r not in lease]
    if missing_sale or missing_lease:
        raise RuntimeError(f"Excel 17개 시도 파싱 실패: 매매={missing_sale}, 전세={missing_lease}")
    return sale, sale_prev, lease, lease_prev


def find_pdf_url(detail_url):
    html = get(detail_url).text
    candidates = []
    for url, text in urls_from_html(html, detail_url):
        blob = f"{url} {text}".lower()
        if ".pdf" in blob or "download" in blob or "file" in blob or "첨부" in text:
            candidates.append(url)
    candidates = list(dict.fromkeys(candidates))
    candidates.sort(key=lambda u: (("pdf" in u.lower()), ("down" in u.lower() or "file" in u.lower())), reverse=True)
    for url in candidates:
        try:
            r = get(url, allow_redirects=True)
        except Exception:
            continue
        if r.content[:4] == b"%PDF" or "application/pdf" in r.headers.get("content-type", "").lower():
            return r.url, r.content
    raise RuntimeError("최신 REB 보도자료에서 PDF 첨부파일을 찾지 못했습니다.")


def pdf_text(pdf_bytes):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join((p.extract_text() or "") for p in reader.pages)
    if len(text) < 500:
        raise RuntimeError("PDF 텍스트 추출량이 너무 적습니다.")
    return text


def appendix_one(text):
    pat = re.compile(r"붙임\s*1\s*[:：]?\s*주간\s*아파트\s*시도별\s*변동률\s*통계표", re.I)
    matches = list(pat.finditer(text))
    if not matches:
        # Some releases insert line breaks or punctuation between words.
        compact = re.sub(r"\s+", "", text)
        marker = "붙임1주간아파트시도별변동률통계표"
        pos = compact.rfind(marker)
        if pos < 0:
            raise RuntimeError("공식 PDF에서 붙임 1 시도별 변동률 통계표를 찾지 못했습니다.")
        # The compact offset cannot safely index the original text; use the last
        # generic '붙임 1' occurrence instead.
        generic = list(re.finditer(r"붙임\s*1", text, re.I))
        if not generic:
            raise RuntimeError("공식 PDF에서 붙임 1을 찾지 못했습니다.")
        start = generic[-1].start()
    else:
        start = matches[-1].start()
    tail = text[start:]
    next_appendix = re.search(r"\n\s*붙임\s*2\b", tail, re.I)
    return tail[:next_appendix.start()] if next_appendix else tail


def split_sale_lease(appendix):
    sale_head = re.search(r"□?\s*매매가격\s*변동률", appendix)
    lease_head = re.search(r"□?\s*전세가격\s*변동률", appendix)
    if not sale_head or not lease_head:
        raise RuntimeError("붙임 1에서 매매/전세 변동률 구간을 찾지 못했습니다.")
    if sale_head.start() < lease_head.start():
        return appendix[sale_head.end():lease_head.start()], appendix[lease_head.end():]
    return appendix[sale_head.end():], appendix[lease_head.end():sale_head.start()]


def region_pattern(region):
    return re.compile(r"^\s*" + r"\s*".join(map(re.escape, region)) + r"(?:\s|$)")


def starts_region(line):
    return any(region_pattern(r).search(line) for r in REGIONS)


def parse_appendix_table(section_text):
    lines = [re.sub(r"[\u00a0\t]+", " ", ln).strip() for ln in section_text.splitlines() if ln.strip()]
    values, previous = {}, {}
    for region in REGIONS:
        rp = region_pattern(region)
        for i, line in enumerate(lines):
            if not rp.search(line):
                continue
            window = line
            nums = re.findall(r"(?<!\d)([-+]?\d+(?:\.\d+)?)(?!\d)", window)
            if len(nums) < 3 and i + 1 < len(lines) and not starts_region(lines[i + 1]):
                window += " " + lines[i + 1]
                nums = re.findall(r"(?<!\d)([-+]?\d+(?:\.\d+)?)(?!\d)", window)
            parsed = [float(x) for x in nums if -100 <= float(x) <= 100]
            if len(parsed) >= 2:
                values[region] = parsed[-1]
                previous[region] = parsed[-2]
                break
    return values, previous


def parse_pdf(detail_url):
    pdf_url, pdf_bytes = find_pdf_url(detail_url)
    appendix = appendix_one(pdf_text(pdf_bytes))
    sale_text, lease_text = split_sale_lease(appendix)
    sale, sale_prev = parse_appendix_table(sale_text)
    lease, lease_prev = parse_appendix_table(lease_text)
    missing_sale = [r for r in REGIONS if r not in sale]
    missing_lease = [r for r in REGIONS if r not in lease]
    if missing_sale or missing_lease:
        raise RuntimeError(f"PDF 붙임표 17개 시도 파싱 실패: 매매={missing_sale}, 전세={missing_lease}")
    return pdf_url, sale, sale_prev, lease, lease_prev


def published_date(detail_html):
    m = re.search(r"등록일\s*</?[^>]*>.*?(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", detail_html, re.S)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", detail_html)
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else datetime.now().strftime("%Y-%m-%d")


def main():
    current = json.loads(OUT.read_text(encoding="utf-8")) if OUT.exists() else {}
    asof_raw, detail_url = latest_release()
    detail_html = get(detail_url).text

    source_method = None
    source_url = None
    sale = sale_prev = lease = lease_prev = None

    try:
        excel_url, excel_bytes = discover_rone_excel(asof_raw)
        sale, sale_prev, lease, lease_prev = parse_excel(excel_bytes, asof_raw)
        source_method = "R-ONE Excel"
        source_url = excel_url
        print("Data source: R-ONE Excel", excel_url)
    except Exception as excel_error:
        print("Excel primary source unavailable; falling back to official PDF:", excel_error, file=sys.stderr)
        pdf_url, sale, sale_prev, lease, lease_prev = parse_pdf(detail_url)
        source_method = "REB official PDF appendix"
        source_url = pdf_url
        print("Data source: official PDF appendix", pdf_url)

    # Hard safety gate: never replace the widget JSON unless all 34 values exist.
    missing = []
    for r in REGIONS:
        if r not in sale: missing.append(f"{r}:sale")
        if r not in lease: missing.append(f"{r}:lease")
    if missing:
        raise RuntimeError("공식 자료 17개 시도 전체 확인 실패: " + ", ".join(missing))

    prev_base = current.get("baseRate", {"value": 2.75, "previous": 2.50, "changedAt": "2026-07-16"})
    data = {
        "source": "Korea Real Estate Board / Bank of Korea",
        "sourceMethod": source_method,
        "sourceUrl": source_url,
        "releaseUrl": detail_url,
        "asOf": f"{asof_raw[:4]}-{asof_raw[4:6]}-{asof_raw[6:]}",
        "published": published_date(detail_html),
        "baseRate": prev_base,
        "regions": [
            {
                "name": r,
                "sale": round(sale[r], 2),
                "salePrev": round(sale_prev[r], 2) if sale_prev.get(r) is not None else None,
                "lease": round(lease[r], 2),
                "leasePrev": round(lease_prev[r], 2) if lease_prev.get(r) is not None else None,
            }
            for r in REGIONS
        ],
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("Latest:", data["asOf"], "published", data["published"])
    print("Updated:", OUT, "via", source_method)


if __name__ == "__main__":
    main()
